from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import uuid
import json
import csv        
import io  
from bson import ObjectId
from cloudinary_config import cloudinary
import cloudinary.uploader
from database import notify_collection,vendors_collection
from utils.email_utils import send_email
 



from database import products_collection

product_bp = Blueprint("products", __name__)



from bson import ObjectId

def normalize_product_id(pid):
    """Ensure product_id is always ObjectId"""
    # case 1: dict like {"$oid": "..."}
    if isinstance(pid, dict) and "$oid" in pid:
        return ObjectId(pid["$oid"])
    # case 2: already ObjectId
    if isinstance(pid, ObjectId):
        return pid
    # case 3: plain string
    if isinstance(pid, str) and ObjectId.is_valid(pid):
        return ObjectId(pid)
    return None



import re
import uuid

def generate_sku(product_name, color="", size=""):
    """
    Smart SKU generator:
    - Includes only available attributes (color, size)
    - Appends short unique suffix
    """
    # Base from product name (max 5 letters, uppercase)
    base = re.sub(r'[^A-Z0-9]', '', product_name.upper())[:5]

    color_part = color.upper()[:3] if color else ""
    size_part = size.upper()[:3] if size else ""
    suffix = str(uuid.uuid4())[:4].upper()  # unique short code

    # Build SKU dynamically
    parts = [base]
    if color_part:
        parts.append(color_part)
    if size_part:
        parts.append(size_part)
    parts.append(suffix)

    return "-".join(parts)

# ------------------ Add Product (Admin or Vendor) ------------------

'''
@product_bp.route('/api/products/add', methods=['POST'])
def add_product():
    try:
        name = request.form['name']
        brand = request.form['brand']
        price = float(request.form['price'])
        discount = float(request.form.get('discount', 0))
        description = request.form['description']
        category = request.form['category']
        sub_category = request.form['subCategory']
        child_category = request.form.get('childCategory', "")
        specifications = json.loads(request.form.get('specifications', '[]'))
        raw_variants = json.loads(request.form.get('variants', '[]'))

        # ✅ Generate SKU for each variant
        variants = []
        for v in raw_variants:
            sku = v.get("sku") or generate_sku(name, v.get("color", ""), v.get("size", ""))
            variants.append({
                "sku": sku,
                "size": v.get("size", ""),
                "color": v.get("color", ""),
                "stock": int(v.get("stock", 0))
            })

        added_by = request.form.get('added_by', 'admin')
        vendor_id = request.form.get('vendor_id')
        status = request.form.get('status', 'active')
        pairs_with_ids = json.loads(request.form.get("pairs_with", "[]"))

        # Handle images
        image_urls = json.loads(request.form.get("images", "[]"))
        files = request.files.getlist('images')
        for file in files:
            if file:
                result = cloudinary.uploader.upload(file)
                image_urls.append(result['secure_url'])

        pairs_with_image_urls = []
        pairs_with_files = request.files.getlist("pairs_with")
        for file in pairs_with_files:
            if file:
                result = cloudinary.uploader.upload(file)
                pairs_with_image_urls.append(result['secure_url'])

        product = {
            "name": name,
            "brand": brand,
            "price": price,
            "discount": discount,
            "description": description,
            "category": category,
            "subCategory": sub_category,
            "childCategory": child_category,
            "specifications": specifications,
            "variants": variants,
            "images": image_urls,
            "pairs_with_images": pairs_with_image_urls,
            "pairs_with": [ObjectId(pid) for pid in pairs_with_ids],
            "created_at": datetime.utcnow(),
            "added_by": added_by,
            "status": status
        }

        if vendor_id:
            product["vendor_id"] = vendor_id

        result = products_collection.insert_one(product)
        product_id = str(result.inserted_id)

        return jsonify({
            "message": "Product added successfully",
            "product_id": product_id
        }), 201

    except Exception as e:
        return jsonify({"error": str(e)}), 500
'''

@product_bp.route('/api/products/add', methods=['POST'])
def add_product():
    try:
        name = request.form['name']
        brand = request.form['brand']
        price = float(request.form['price'])
        discount = float(request.form.get('discount', 0))
        description = request.form['description']
        category = request.form['category']
        sub_category = request.form['subCategory']
        child_category = request.form.get('childCategory', "")
        specifications = json.loads(request.form.get('specifications', '[]'))
        raw_variants = json.loads(request.form.get('variants', '[]'))

        # ✅ Generate SKU for each variant
        variants = []
        for v in raw_variants:
            sku = v.get("sku") or generate_sku(name, v.get("colorName", v.get("color", "")), v.get("size", ""))
            variant = {
                "sku": sku,
                "size": v.get("size", ""),
                "color": v.get("colorName", v.get("color", "")),      # backward compat
                "colorName": v.get("colorName", v.get("color", "")),  # "Navy Blue"
                "colorHex": v.get("colorHex", ""),                    # "#1a237e"
                "stock": int(v.get("stock", 0)),
            }
            # Only save measurements if at least one field is filled
            '''
            measurements = {
                "chest":    str(v.get("chest", "")),
                "waist":    str(v.get("waist", "")),
                "hips":     str(v.get("hips", "")),
                "shoulder": str(v.get("shoulder", "")),
                "length":   str(v.get("length", "")),
            }
            '''
            measurement_data = v.get("measurements", {})

            measurements = {
    "chest":    str(measurement_data.get("chest", "")),
    "waist":    str(measurement_data.get("waist", "")),
    "hips":     str(measurement_data.get("hips", "")),
    "shoulder": str(measurement_data.get("shoulder", "")),
    "length":   str(measurement_data.get("length", "")),
}
            if any(measurements.values()):
                variant["measurements"] = measurements
            variants.append(variant)

        added_by = request.form.get('added_by', 'admin')
        vendor_id = request.form.get('vendor_id')
        status = request.form.get('status', 'active')
        pairs_with_ids = json.loads(request.form.get("pairs_with", "[]"))

        # Handle images
        image_urls = json.loads(request.form.get("images", "[]"))
        files = request.files.getlist('images')
        for file in files:
            if file:
                result = cloudinary.uploader.upload(file)
                image_urls.append(result['secure_url'])

        pairs_with_image_urls = []
        pairs_with_files = request.files.getlist("pairs_with")
        for file in pairs_with_files:
            if file:
                result = cloudinary.uploader.upload(file)
                pairs_with_image_urls.append(result['secure_url'])

        product = {
            "name": name,
            "brand": brand,
            "price": price,
            "discount": discount,
            "description": description,
            "category": category,
            "subCategory": sub_category,
            "childCategory": child_category,
            "specifications": specifications,
            "variants": variants,
            "images": image_urls,
            "pairs_with_images": pairs_with_image_urls,
            "pairs_with": [ObjectId(pid) for pid in pairs_with_ids],
            "created_at": datetime.utcnow(),
            "added_by": added_by,
            "status": status
        }

        if vendor_id:
            product["vendor_id"] = vendor_id

        result = products_collection.insert_one(product)
        product_id = str(result.inserted_id)

        return jsonify({
            "message": "Product added successfully",
            "product_id": product_id
        }), 201

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ------------------ Get All Products ------------------
'''
@product_bp.route('/api/products/all', methods=['GET'])
def get_all_products():
    try:
        # Detect admin mode (?admin=true)
        is_admin = request.args.get("admin", "").lower() == "true"

        # Admin sees everything, customers see only approved/active
        if is_admin:
            query = {}
        else:
            query = {
                "$or": [
                    # Admin-added products visible to customers
                    {"added_by": "admin", "status": {"$in": ["active", "approved"]}},
                    # Vendor products visible only after approval
                    {"added_by": "vendor", "status": "approved"}
                ]
            }
       
        products = list(products_collection.find(query).sort("created_at", -1))

        for p in products:
            p["_id"] = str(p["_id"])
            p = normalize_product(p)

            # Convert ObjectIds in pairs_with to strings
            if "pairs_with" in p and isinstance(p["pairs_with"], list):
                p["pairs_with"] = [str(pid) for pid in p["pairs_with"]]

        return jsonify({"products": products}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
'''
@product_bp.route('/api/products/all', methods=['GET'])
def get_all_products():
    try:
        is_admin = request.args.get("admin", "").lower() == "true"

        if is_admin:
            query = {}
        else:
            query = {
                "$or": [
                    {"added_by": "admin",  "status": {"$in": ["active", "approved"]}},
                    {"added_by": "vendor", "status": "approved"},   # ✅ only approved
                ]
            }

        raw = list(products_collection.find(query).sort("created_at", -1))

        products = []
        for p in raw:
            p["_id"] = str(p["_id"])
            p = normalize_product(p)

            if "pairs_with" in p and isinstance(p["pairs_with"], list):
                p["pairs_with"] = [str(pid) for pid in p["pairs_with"]]

            products.append(p)

        return jsonify({"products": products}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def normalize_product(product, auto_save=True):
    
    updated = False
    for variant in product.get("variants", []):
        # Normalize stock
        stock_val = variant.get("stock", {"$numberInt": "0"})
        if isinstance(stock_val, dict) and "$numberInt" in stock_val:
            variant["stock"] = int(stock_val["$numberInt"])
        else:
            variant["stock"] = int(stock_val)

        # Generate SKU if missing
        if not variant.get("sku"):
            variant["sku"] = generate_sku(
                product.get("name", "PROD"),
                variant.get("color", ""),
                variant.get("size", "")
            )
            updated = True

    # Normalize category keys
    product["subcategory"] = product.get("subcategory") or product.get("subCategory")
    product["childcategory"] = product.get("childcategory") or product.get("childCategory")

    # Auto-save new SKUs to MongoDB
    if auto_save and updated and "_id" in product:
        products_collection.update_one(
            {"_id": product["_id"]},
            {"$set": {"variants": product["variants"]}}
        )

    return product
# ------------------ Get Single Product ------------------

import traceback
'''
@product_bp.route('/api/products/<product_id>', methods=['GET'])
def get_product_by_id(product_id):
    try:
        product = products_collection.find_one({"_id": ObjectId(product_id)})
        if not product:
            return jsonify({"error": "Product not found"}), 404

        product["_id"] = str(product["_id"])
        product = normalize_product(product)

        # Convert ObjectId list for pairs_with
        if "pairs_with" in product and isinstance(product["pairs_with"], list):
            product["pairs_with"] = [str(pid) for pid in product["pairs_with"]]

            # Fetch full paired product details
            paired_products = []
            for pid in product["pairs_with"]:
                paired = products_collection.find_one({"_id": ObjectId(pid)}, {
                    "_id": 1,
                    "name": 1,
                    "price": 1,
                    "discount": 1,
                    "images": 1
                })
                if paired:
                    paired["_id"] = str(paired["_id"])
                    
                    # Ensure at least one image is available
                    paired["image"] = paired["images"][0] if paired.get("images") else None
                    
                    # Calculate discounted price
                    original_price = float(paired.get("price", 0))
                    discount = float(paired.get("discount", 0))
                    discounted_price = original_price - (original_price * discount / 100)
                    paired["final_price"] = round(discounted_price, 2)
                    
                    paired_products.append(paired)

            product["pairs_with_products"] = paired_products

        return jsonify({"product": product}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
     

def get_product_with_pairs(product):
    product = dict(product)  
    pair_ids = product.get("pairs_with", [])
    
    if pair_ids:
        product["pairs_with_products"] = list(products_collection.find(
            {"_id": {"$in": pair_ids}},
            {"name": 1, "price": 1, "discount": 1, "images": 1}
        ))
        
        # Add final_price for each paired product
        for p in product["pairs_with_products"]:
            p["final_price"] = round(
                p["price"] - (p.get("discount", 0) / 100 * p["price"])
            )
            if p.get("images"):
                p["image"] = p["images"][0]
    else:
        product["pairs_with_products"] = []
    
    return product
'''


@product_bp.route('/api/products/<product_id>', methods=['GET'])
def get_product_by_id(product_id):
    try:
        product = products_collection.find_one({"_id": ObjectId(product_id)})
        if not product:
            return jsonify({"error": "Product not found"}), 404

        product["_id"] = str(product["_id"])
        product = normalize_product(product)

        # Convert ObjectIds in pairs_with to strings
        if "pairs_with" in product and isinstance(product["pairs_with"], list):
            product["pairs_with"] = [str(pid) for pid in product["pairs_with"]]

            paired_products = []
            for pid in product["pairs_with"]:
                paired = products_collection.find_one({"_id": ObjectId(pid)}, {
                    "_id": 1, "name": 1, "price": 1, "discount": 1, "images": 1
                })
                if paired:
                    paired["_id"] = str(paired["_id"])
                    paired["image"] = paired["images"][0] if paired.get("images") else None
                    original_price = float(paired.get("price", 0))
                    discount = float(paired.get("discount", 0))
                    paired["final_price"] = round(original_price - (original_price * discount / 100), 2)
                    paired_products.append(paired)
            product["pairs_with_products"] = paired_products

        # ✅ BUILD variant map: tells frontend which colors exist per size
        # and which sizes exist per color — so invalid combos are disabled
        variants = product.get("variants", [])

        # colors available for each size
        size_to_colors = {}
        # sizes available for each color
        color_to_sizes = {}

        for v in variants:
            s = v.get("size", "").strip()
            c = v.get("color", "").strip()
            stk = v.get("stock", 0)
            if isinstance(stk, dict):
                stk = int(stk.get("$numberInt", 0))
            stk = int(stk)

            if s:
                if s not in size_to_colors:
                    size_to_colors[s] = []
                if c:
                    size_to_colors[s].append({
                        "color": c,
                        "stock": stk,
                        "inStock": stk > 0
                    })

            if c:
                if c not in color_to_sizes:
                    color_to_sizes[c] = []
                if s:
                    color_to_sizes[c].append({
                        "size": s,
                        "stock": stk,
                        "inStock": stk > 0
                    })

        product["size_to_colors"] = size_to_colors   # { "S": [{color,stock,inStock},...] }
        product["color_to_sizes"] = color_to_sizes   # { "green": [{size,stock,inStock},...] }

        return jsonify({"product": product}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



@product_bp.route('/api/products/<product_id>/sizechart', methods=['GET'])
def get_size_chart(product_id):
    try:
        product = products_collection.find_one({"_id": ObjectId(product_id)})
        if not product:
            return jsonify({"error": "Product not found"}), 404

        variants = product.get("variants", [])

        # Build size chart: one row per unique size
        size_chart = {}
        for v in variants:
            s = v.get("size", "").strip()
            if s and s not in size_chart:
                m = v.get("measurements", {})
                if m:  # only include if measurements exist
                    size_chart[s] = {
                        "chest":    m.get("chest", ""),
                        "waist":    m.get("waist", ""),
                        "hips":     m.get("hips", ""),
                        "shoulder": m.get("shoulder", ""),
                        "length":   m.get("length", ""),
                    }

        return jsonify({
            "product_id": product_id,
            "category":   product.get("category", ""),
            "size_chart": size_chart
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
'''

@product_bp.route('/api/products/<product_id>/sizechart/upload', methods=['POST'])
def upload_size_chart(product_id):
    try:
        product = products_collection.find_one({"_id": ObjectId(product_id)})
        if not product:
            return jsonify({"error": "Product not found"}), 404

        # Option A — JSON body  {"S": {"chest": 30, "waist": 26}, "M": {...}}
        if request.is_json:
            size_chart_data = request.get_json()

        # Option B — CSV file upload
        elif 'file' in request.files:
            file = request.files['file']
            stream = io.StringIO(file.stream.read().decode("utf-8"))
            reader = csv.DictReader(stream)
            size_chart_data = {}
            for row in reader:
                size = row.pop("size", "").strip()   # CSV must have a "size" column
                if size:
                    size_chart_data[size] = {
                        k.strip(): v.strip() for k, v in row.items() if v.strip()
                    }
        else:
            return jsonify({"error": "Send JSON body or a CSV file"}), 400

        # Update each matching variant's measurements
        variants = product.get("variants", [])
        for variant in variants:
            s = variant.get("size", "").strip()
            if s and s in size_chart_data:
                variant["measurements"] = size_chart_data[s]

        products_collection.update_one(
            {"_id": ObjectId(product_id)},
            {"$set": {"variants": variants}}
        )

        return jsonify({
            "message": "Size chart uploaded successfully",
            "sizes_updated": list(size_chart_data.keys())
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

'''
# Replace the existing upload_size_chart route in product_routes.py

import io, csv
import cloudinary.uploader

SIZECHART_ALLOWED = {"png", "jpg", "jpeg", "webp", "pdf", "csv", "xlsx", "xls"}

def allowed_sizechart(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in SIZECHART_ALLOWED


@product_bp.route('/api/products/<product_id>/sizechart/upload', methods=['POST'])
def upload_size_chart(product_id):
    try:
        product = products_collection.find_one({"_id": ObjectId(product_id)})
        if not product:
            return jsonify({"error": "Product not found"}), 404

        # ── Auth check: admin passes ?is_admin=true, vendor passes their vendor_id ──
        is_admin   = request.args.get("is_admin", "").lower() == "true" \
                     or request.form.get("is_admin", "").lower() == "true"
        vendor_id  = request.args.get("vendor_id") or request.form.get("vendor_id")

        if is_admin:
            pass  # admin can upload for any product
        elif vendor_id:
            # vendor can only upload for their own product
            if product.get("vendor_id") != vendor_id:
                return jsonify({"error": "Access denied — not your product"}), 403
        else:
            return jsonify({"error": "Unauthorized — pass is_admin=true or vendor_id"}), 401

        # ── Option A: JSON body {"S": {"chest": 30, ...}, "M": {...}} ──────────
        if request.is_json:
            size_chart_data = request.get_json()
            _apply_measurements(product, size_chart_data, product_id)
            return jsonify({
                "message": "Size chart saved",
                "sizes_updated": list(size_chart_data.keys()),
            }), 200

        # ── Option B: file upload ──────────────────────────────────────────────
        if "file" not in request.files:
            return jsonify({"error": "Send JSON body or a file"}), 400

        file     = request.files["file"]
        filename = file.filename or ""
        ext      = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        if not allowed_sizechart(filename):
            return jsonify({"error": f"File type '{ext}' not allowed"}), 400

        # ── CSV ────────────────────────────────────────────────────────────────
        if ext == "csv":
            stream = io.StringIO(file.stream.read().decode("utf-8"))
            reader = csv.DictReader(stream)
            size_chart_data = {}
            for row in reader:
                size = row.pop("size", "").strip()
                if size:
                    size_chart_data[size] = {
                        k.strip(): v.strip() for k, v in row.items() if v.strip()
                    }
            _apply_measurements(product, size_chart_data, product_id)
            return jsonify({
                "message": "Size chart (CSV) uploaded and parsed",
                "sizes_updated": list(size_chart_data.keys()),
            }), 200

        # ── Excel ──────────────────────────────────────────────────────────────
        if ext in ("xlsx", "xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file.stream)
                ws = wb.active
                headers = [str(cell.value).strip().lower() for cell in ws[1]]
                size_chart_data = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {
                        headers[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row)
                    }
                    size = row_dict.pop("size", "").strip()
                    if size:
                        size_chart_data[size] = {k: v for k, v in row_dict.items() if v}
                _apply_measurements(product, size_chart_data, product_id)
                return jsonify({
                    "message": "Size chart (Excel) uploaded and parsed",
                    "sizes_updated": list(size_chart_data.keys()),
                }), 200
            except ImportError:
                return jsonify({"error": "openpyxl not installed — run: pip install openpyxl"}), 500
            except Exception as e:
                return jsonify({"error": f"Excel parse error: {str(e)}"}), 400

        # ── Image or PDF → Cloudinary ──────────────────────────────────────────
        if ext in ("png", "jpg", "jpeg", "webp", "pdf"):
            resource_type = "raw" if ext == "pdf" else "image"
            upload_result = cloudinary.uploader.upload(
                file,
                folder=f"citimart/sizecharts/{product_id}",
                resource_type=resource_type,
                public_id=f"sizechart_{product_id}",
                overwrite=True,
            )
            url = upload_result["secure_url"]
            products_collection.update_one(
                {"_id": ObjectId(product_id)},
                {"$set": {"size_chart_url": url, "size_chart_type": ext}}
            )
            return jsonify({
                "message": f"Size chart ({ext.upper()}) uploaded successfully",
                "url": url,
            }), 200

        return jsonify({"error": "Unsupported file type"}), 400

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _apply_measurements(product, size_chart_data, product_id):
    """Apply parsed size chart rows to matching variant measurements."""
    variants = product.get("variants", [])
    for variant in variants:
        s = variant.get("size", "").strip()
        if s and s in size_chart_data:
            variant["measurements"] = size_chart_data[s]
    products_collection.update_one(
        {"_id": ObjectId(product_id)},
        {"$set": {"variants": variants}}
    )


# ------------------ Update Product ------------------


from flask import request, jsonify
from bson import ObjectId
from datetime import datetime
import json, os
from database import products_collection, notify_collection
from utils.email_utils import send_email  
import cloudinary
'''
@product_bp.route('/api/products/<product_id>', methods=['PUT'])
def update_product(product_id):
    try:
        vendor_id = request.form.get('vendor_id')
        is_admin = request.form.get('is_admin') == 'true'

        query = {'_id': ObjectId(product_id)}
        if vendor_id and not is_admin:
            query.update({'vendor_id': vendor_id, 'added_by': 'vendor'})
        elif is_admin:
            query.update({'added_by': 'admin'})
        else:
            return jsonify({'error': 'Access denied'}), 403

        existing_product = products_collection.find_one(query)
        if not existing_product:
            return jsonify({'error': 'Product not found'}), 404

        # --- Name and Variants ---
        name = request.form.get('name', existing_product['name'])
        raw_variants = json.loads(request.form.get('variants', '[]'))

        # Generate SKU for missing variants
        variants = []
        for v in raw_variants:
            sku = v.get("sku") or generate_sku(name, v.get("color", ""), v.get("size", ""))
            variants.append({
                "sku": sku,
                "size": v.get("size", ""),
                "color": v.get("color", ""),
                "stock": int(v.get("stock", 0))
            })

        # --- Update fields ---
        update_fields = {
            "name": name,
            "description": request.form.get('description', existing_product.get('description', '')),
            "price": float(request.form.get('price', existing_product.get('price', 0))),
            "category": request.form.get('category', existing_product.get('category', '')),
            "subCategory": request.form.get('subCategory', existing_product.get('subCategory', '')),
            "childCategory": request.form.get('childCategory', existing_product.get('childCategory', '')),
            "brand": request.form.get('brand', existing_product.get('brand', '')),
            "discount": float(request.form.get('discount', existing_product.get('discount', 0))),
            "status": request.form.get('status', existing_product.get('status', 'active')),
            "specifications": json.loads(request.form.get('specifications', '[]')),
            "variants": variants,
            "updated_at": datetime.utcnow()
        }

        # --- Handle pairs_with ---
        pairs_with_ids = json.loads(request.form.get("pairs_with", "[]"))
        update_fields["pairs_with"] = [ObjectId(pid) for pid in pairs_with_ids]

        # --- Handle images ---
        image_urls = request.form.getlist("images")
        files = request.files.getlist("images")
        for file in files:
            if file:
                result = cloudinary.uploader.upload(file)
                image_urls.append(result['secure_url'])

        if not image_urls:
            image_urls = existing_product.get("images", [])
        update_fields["images"] = image_urls

        # --- Update in DB ---
        result = products_collection.update_one(query, {"$set": update_fields})

        # --- Notify customers if product is back in stock ---
        # Only send email if at least one variant has stock > 0
        if any(v.get("stock", 0) > 0 for v in variants):
            subscribers = list(notify_collection.find({"product_id": str(product_id)}))
            for sub in subscribers:
                subject = "Good news! Your product is back in stock 🎉"
                body = f"""
Hi,

The product '{update_fields.get('name', 'a product')}' you requested is now available again!

👉 Check it here: {os.getenv("FRONTEND_URL")}/products/{product_id}

Thank you,
Citimart Team
"""
                send_email(sub["customer_email"], subject, body)

            # Remove subscribers after sending emails
            notify_collection.delete_many({"product_id": str(product_id)})

        return jsonify({'message': 'Product updated successfully'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
'''

@product_bp.route('/api/products/<product_id>', methods=['PUT'])
def update_product(product_id):
    try:
        vendor_id = request.form.get('vendor_id')
        is_admin = request.form.get('is_admin') == 'true'

        query = {'_id': ObjectId(product_id)}
        if vendor_id and not is_admin:
            query.update({'vendor_id': vendor_id, 'added_by': 'vendor'})
        elif is_admin:
            query.update({'added_by': 'admin'})
        else:
            return jsonify({'error': 'Access denied'}), 403

        existing_product = products_collection.find_one(query)
        if not existing_product:
            return jsonify({'error': 'Product not found'}), 404

        # --- Name and Variants ---
        name = request.form.get('name', existing_product['name'])
        raw_variants = json.loads(request.form.get('variants', '[]'))

        # Generate SKU for missing variants
        variants = []
        for v in raw_variants:
            sku = v.get("sku") or generate_sku(name, v.get("colorName", v.get("color", "")), v.get("size", ""))
            variant = {
                "sku": sku,
                "size": v.get("size", ""),
                "color": v.get("colorName", v.get("color", "")),      # backward compat
                "colorName": v.get("colorName", v.get("color", "")),  # "Navy Blue"
                "colorHex": v.get("colorHex", ""),                    # "#1a237e"
                "stock": int(v.get("stock", 0)),
            }
            # Only save measurements if at least one field is filled
            m = v.get("measurements", {})   
            measurements = {
                "chest":    str(m.get("chest",    "")),
                "waist":    str(m.get("waist",    "")),
                "hips":     str(m.get("hips",     "")),
                "shoulder": str(m.get("shoulder", "")),
                "length":   str(m.get("length",   "")),
            }

            if any(measurements.values()):
                variant["measurements"] = measurements
            variants.append(variant)

        # --- Update fields ---
        update_fields = {
            "name": name,
            "description": request.form.get('description', existing_product.get('description', '')),
            "price": float(request.form.get('price', existing_product.get('price', 0))),
            "category": request.form.get('category', existing_product.get('category', '')),
            "subCategory": request.form.get('subCategory', existing_product.get('subCategory', '')),
            "childCategory": request.form.get('childCategory', existing_product.get('childCategory', '')),
            "brand": request.form.get('brand', existing_product.get('brand', '')),
            "discount": float(request.form.get('discount', existing_product.get('discount', 0))),
            "status": request.form.get('status', existing_product.get('status', 'active')),
            "specifications": json.loads(request.form.get('specifications', '[]')),
            "variants": variants,
            "updated_at": datetime.utcnow()
        }

        # --- Handle pairs_with ---
        pairs_with_ids = json.loads(request.form.get("pairs_with", "[]"))
        update_fields["pairs_with"] = [ObjectId(pid) for pid in pairs_with_ids]

        # --- Handle images ---
        image_urls = request.form.getlist("images")
        files = request.files.getlist("images")
        for file in files:
            if file:
                result = cloudinary.uploader.upload(file)
                image_urls.append(result['secure_url'])

        if not image_urls:
            image_urls = existing_product.get("images", [])
        update_fields["images"] = image_urls

        # --- Update in DB ---
        result = products_collection.update_one(query, {"$set": update_fields})

        # --- Notify customers if product is back in stock ---
        if any(v.get("stock", 0) > 0 for v in variants):
            subscribers = list(notify_collection.find({"product_id": str(product_id)}))
            for sub in subscribers:
                subject = "Good news! Your product is back in stock 🎉"
                body = f"""
Hi,

The product '{update_fields.get('name', 'a product')}' you requested is now available again!

👉 Check it here: {os.getenv("FRONTEND_URL")}/products/{product_id}

Thank you,
Citimart Team
"""
                send_email(sub["customer_email"], subject, body)

            notify_collection.delete_many({"product_id": str(product_id)})

        return jsonify({'message': 'Product updated successfully'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ------------------ Get Full Products by Category/Subcategory/ChildCategory ------------------


@product_bp.route('/api/products/by-category', methods=['GET'])
def get_products_by_category():
    try:
        # Get list values instead of single strings
        categories = request.args.getlist('category')
        subcategories = request.args.getlist('subcategory')
        child_categories = request.args.getlist('child_category')

        query = {}
        if categories:
            query["category"] = {"$in": categories}
        if subcategories:
            query["$or"] = [
        {"subCategory": {"$in": subcategories}},
        {"subcategory": {"$in": subcategories}}
          ]
        if child_categories:
           query.setdefault("$and", []).append({
        "$or": [
            {"childCategory": {"$in": child_categories}},
            {"childcategory": {"$in": child_categories}}
        ]
    })


        products_cursor = products_collection.find(query).sort("created_at", -1)

        product_list = []
        for product in products_cursor:
            product['_id'] = str(product['_id'])

            # Resolve pairs_with to product details
            if "pairs_with" in product and isinstance(product["pairs_with"], list):
                paired_ids = [ObjectId(pid) for pid in product["pairs_with"] if ObjectId.is_valid(pid)]
                paired_products_cursor = products_collection.find(
                    {"_id": {"$in": paired_ids}},
                    {"_id": 1, "name": 1, "price": 1, "discount": 1, "stock": 1, "images": 1}
                )
                paired_products = [{
                    "id": str(p["_id"]),
                    "name": p.get("name"),
                    "price": p.get("price"),
                    "discount": p.get("discount"),
                    "stock": p.get("stock"),
                    "images": p.get("images", [])
                } for p in paired_products_cursor]
                product["pairs_with"] = paired_products
            else:
                product["pairs_with"] = []

            product_list.append(product)

        return jsonify({"products": product_list}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


#----- Frequently bought together section --------------

from flask import jsonify, request, Response
from bson import ObjectId, json_util
import json

@product_bp.route('/api/products/frequently-bought/<product_id>', methods=['GET'])
def get_frequently_bought(product_id):
    try:
        # 🔒 Validate product_id
        if not ObjectId.is_valid(product_id):
            return jsonify({"error": "Invalid product ID"}), 400

        # 🔎 Find the main product
        product = products_collection.find_one({"_id": ObjectId(product_id)})
        if not product:
            return jsonify({"error": "Product not found"}), 404

        # 🧾 Build query for related products
        query = {
            "category": product.get("category", ""),
            "subCategory": product.get("subCategory", ""),
            "_id": {"$ne": ObjectId(product_id)}
        }

        if product.get("childCategory"):
            query["childCategory"] = product["childCategory"]

        # 📦 Fetch related products
        related = list(products_collection.find(query).limit(3))

        # ✅ Convert BSON → JSON-safe
        return Response(
            response=json.dumps({"relatedProducts": related}, default=json_util.default),
            status=200,
            mimetype="application/json"
        )

    except Exception as e:
        print("Error in get_frequently_bought:", str(e))
        return jsonify({"error": "Internal server error"}), 500


# ------------------ Delete Product ------------------
@product_bp.route('/api/products/<product_id>', methods=['DELETE'])
def delete_product(product_id):
    try:
        vendor_id = request.args.get('vendor_id')
        is_admin = request.args.get('is_admin') == 'true'

        query = {'_id': ObjectId(product_id)}
        if vendor_id and not is_admin:
            query.update({'vendor_id': vendor_id, 'added_by': 'vendor'})
        elif is_admin:
            query.update({'added_by': 'admin'})
        else:
            return jsonify({'error': 'Access denied'}), 403

        result = products_collection.delete_one(query)

        if result.deleted_count == 1:
            return jsonify({'message': 'Product deleted successfully'}), 200
        else:
            print(f"[DEBUG] Delete failed. Query used: {query}")
            return jsonify({'error': 'Product not found or access denied'}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ------------------ Get All Products by Vendor ------------------
@product_bp.route('/api/products/vendor/<vendor_id>', methods=['GET'])
def get_products_by_vendor(vendor_id):
    try:
        products = list(products_collection.find({
            "vendor_id": vendor_id,
            "added_by": "vendor"
        }).sort("created_at", -1))
        for p in products:
            p['_id'] = str(p['_id'])
        return jsonify({"products": products}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ------------------ Get Product IDs by Category/Subcategory/ChildCategory ------------------
@product_bp.route('/api/products/ids-by-category', methods=['GET'])
def get_product_ids_by_category():
    try:
        category = request.args.get('category')
        subcategory = request.args.get('subcategory')
        child_category = request.args.get('child_category')

        query = {}
        if category:
            query["category"] = category
        if subcategory:
            query["$or"] = [
        {"subCategory": subcategory},
        {"subcategory": subcategory}
    ]
        if child_category:
            query.setdefault("$and", []).append({
        "$or": [
            {"childCategory": child_category},
            {"childcategory": child_category}
        ]
    })


        products = products_collection.find(query, {"_id": 1, "name": 1})
        product_ids = [{"id": str(product["_id"]), "name": product["name"]} for product in products]

        return jsonify({"product_ids": product_ids}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@product_bp.route('/api/products', methods=['GET'])
def get_products():
    try:
        products = list(products_collection.find().sort("created_at", -1))
        for p in products:
            p['_id'] = str(p['_id'])
            if "pairs_with" in p and isinstance(p["pairs_with"], list):
                p["pairs_with"] = [str(pid) for pid in p["pairs_with"]]
        return jsonify({"products": products}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@product_bp.route('/api/products/similar/<product_id>', methods=['GET'])
def get_similar_products(product_id):
    try:
        product = products_collection.find_one({'_id': ObjectId(product_id)})
        if not product:
            return jsonify({'error': 'Product not found'}), 404

        query = {
            "category": product.get("category", ""),
            "_id": {"$ne": ObjectId(product_id)}
        }

        similar = list(products_collection.find(query).limit(4))

        for p in similar:
            p['_id'] = str(p['_id'])

        return jsonify({'products': similar}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

'''
@product_bp.route('/products/new-arrivals', methods=['GET'])
def get_new_arrivals():
    try:
        products = list(products_collection.find(
            #{"status": "active"}  # only active products
            {"status": {"$in": ["active", "approved"]}}
        ).sort("created_at", -1).limit(8))  # latest 8 products

        # Convert ObjectId and other non-JSON types
        for p in products:
            p["_id"] = str(p["_id"])
            p["price"] = float(p.get("price", 0))
            p["discount"] = float(p.get("discount", 0))
            p["created_at"] = p["created_at"]

        return jsonify(products), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
'''
@product_bp.route('/products/new-arrivals', methods=['GET'])
def get_new_arrivals():
    try:
        products = list(
            products_collection.find(
                {"status": {"$in": ["active", "approved"]}}
            )
            .sort("created_at", -1)
            .limit(8)
        )

        result = []

        for p in products:
            product = dict(p)

            # Convert Mongo ObjectId
            product["_id"] = str(product["_id"])

            # Safe price/discount conversion
            try:
                product["price"] = float(product.get("price") or 0)
            except:
                product["price"] = 0

            try:
                product["discount"] = float(product.get("discount") or 0)
            except:
                product["discount"] = 0

            # Convert datetime fields
            for field in [
                "created_at",
                "approved_at",
                "reviewed_at",
                "updated_at"
            ]:
                if field in product and isinstance(product[field], datetime):
                    product[field] = product[field].isoformat()

            # Convert pairs_with ObjectIds
            if "pairs_with" in product and isinstance(product["pairs_with"], list):
                product["pairs_with"] = [str(pid) for pid in product["pairs_with"]]

            result.append(product)

        return jsonify(result), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "error": str(e)
        }), 500    


@product_bp.route("/api/products/review/<product_id>", methods=["PATCH"])
def review_product(product_id):
    try:
        data = request.get_json()
        new_status = data.get("status")
        admin_comment = data.get("admin_comment", "").strip()

        # --- Validate new status ---
        allowed_statuses = ["Suspended", "Rejected", "Hidden", "Under Review", "Approved"]
        if new_status not in allowed_statuses:
            return jsonify({"error": "Invalid status"}), 400

        # --- Find product in DB ---
        product = products_collection.find_one({"_id": ObjectId(product_id)})
        if not product:
            return jsonify({"error": "Product not found"}), 404

        # --- Restrict moderation to vendor-added products only ---
        if product.get("added_by") != "vendor":
            return jsonify({"error": "Only vendor-added products can be moderated"}), 403

        # --- Prepare update fields ---
        update_fields = {
            "status": new_status.lower(),  # keep status consistent (lowercase)
            "admin_comment": admin_comment,
            "reviewed_at": datetime.utcnow(),
        }

        # --- If status is Approved, mark product as visible ---
        if new_status.lower() == "approved":
            update_fields["visible_on_site"] = True  # optional flag for future use
            update_fields["approved_at"] = datetime.utcnow()

        # --- Push to review history ---
        review_entry = {
            "action": new_status,
            "by": "admin",
            "comment": admin_comment,
            "timestamp": datetime.utcnow(),
        }

        # --- Update DB ---
        products_collection.update_one(
            {"_id": ObjectId(product_id)},
            {
                "$set": update_fields,
                "$push": {"review_history": review_entry},
            },
        )

        # --- Optional: Notify vendor (placeholder) ---
        vendor_id = product.get("vendor_id")
        if vendor_id:
            vendor = vendors_collection.find_one({"_id": ObjectId(vendor_id)})
            if vendor and vendor.get("email"):
                print(
                    f"📩 Vendor {vendor['email']} notified: "
                    f"Product '{product.get('name')}' status changed → {new_status}"
                )

        # --- Response ---
        return jsonify({
            "message": f"✅ Product status updated to '{new_status}'",
            "product_id": str(product_id),
            "status": new_status.lower(),
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

#http://localhost:5000/api/products/ids-by-category?category=Clothing&subcategory=Women&child_category=Tops

